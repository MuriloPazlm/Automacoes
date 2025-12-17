import os
import re
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# Caminhos
pasta_base = r''
arquivo_origem = r''

# Regras de ciclos
CICLOS = [
    {"nome": "CICLO 07", "inicio": "2025-07-18", "fim": "2025-08-15"},
    {"nome": "CICLO 08", "inicio": "2025-08-18", "fim": "2025-10-03"},
    {"nome": "CICLO 09", "inicio": "2025-09-16", "fim": "2025-11-14"},
    {"nome": "CICLO 10", "inicio": "2025-11-17", "fim": "2025-12-12"},
]
for ciclo in CICLOS:
    ciclo["inicio"] = datetime.strptime(ciclo["inicio"], "%Y-%m-%d")
    ciclo["fim"] = datetime.strptime(ciclo["fim"], "%Y-%m-%d")

def obter_ciclo_por_data(data):
    for ciclo in CICLOS:
        if ciclo["inicio"] <= data <= ciclo["fim"]:
            return ciclo["nome"]
    return None

# Carrega a planilha
wb = load_workbook(arquivo_origem)
ws = wb.active
colunaBrick = 7

# Etapa 1: Identificar o número máximo de BRICKs

max_splits = 1
for row in range(2, ws.max_row + 1):
    brick = ws.cell(row=row, column=colunaBrick).value
    if brick and isinstance(brick, str):
        partes = re.split(r'\s+', brick.strip())
        if len(partes) > max_splits:
            max_splits = len(partes)

# Etapa 2: Inserir colunas extras
for _ in range(max_splits - 1):
    ws.insert_cols(colunaBrick + 1)


# Etapa 3: Separar os BRICKs

for row in range(2, ws.max_row + 1):
    brick = ws.cell(row=row, column=colunaBrick).value
    if brick and isinstance(brick, str):
        partes = re.split(r'\s+', brick.strip())
        for i, parte in enumerate(partes):
            ws.cell(row=row, column=colunaBrick + i, value=parte)


# Etapa 4: Inserir colunas para VLOOKUP
colunaFinal = colunaBrick + max_splits
ws.insert_cols(colunaFinal)
ws.insert_cols(colunaFinal + 1)

setorRep = input("Informe o Setor do Representante: ")
baseBrick_input = input("Cole os códigos BRICK separados por espaço, vírgula ou quebra de linha:\n")
codigosRaw = re.split(r'[,\s]+', baseBrick_input.strip())

def format(codigo):
    codigo = codigo.strip()
    if codigo.startswith("BR_") and len(codigo) == 10:
        return f'{codigo}'
    else:
        return f'BR_{codigo.zfill(7)}'

baseBrick = [format(c) for c in codigosRaw if c]

for i, codigo in enumerate(baseBrick, start=2):
    ws.cell(row=i, column=colunaFinal + 1, value=codigo)

# Etapa 5: Fórmulas VLOOKUP
coluna_vazia = colunaFinal
coluna_codigos_usuario = colunaFinal + 1
ultima_linha = ws.max_row

for row in range(2, ultima_linha + 1):
    letra_coluna_g = get_column_letter(colunaBrick)
    letra_coluna_usuario = get_column_letter(coluna_codigos_usuario)
    letra_coluna_vazia = get_column_letter(coluna_vazia)
    formula = f'=IFERROR(VLOOKUP({letra_coluna_g}{row},{letra_coluna_usuario}:{letra_coluna_usuario},1,FALSE),"")'
    ws[f'{letra_coluna_vazia}{row}'] = formula

# Etapa 6: Criar aba Limpeza
if "Limpeza" not in wb.sheetnames:
    ws_Limpeza = wb.create_sheet("Limpeza")
else:
    ws_Limpeza = wb["Limpeza"]
    ws_Limpeza.delete_rows(1, ws_Limpeza.max_row)

cabecalhos = {
    "Lista de clientes-alvo: Nome": None,
    "Alvo: Alvos": None,
    "Nome da conta": None,
    "Ciclo de Marketing": None
}

for col in range(1, ws.max_column + 1):
    valor = ws.cell(row=1, column=col).value
    if valor in cabecalhos:
        cabecalhos[valor] = col

ws_Limpeza.append([
    "Lista de clientes-alvo:",
    "Primeiro Nome (Alvo)", "Sobrenomes (Alvo)",
    "Nome da conta", "Total Calls", "Ciclo de Marketing"
])

for row in range(2, ultima_linha + 1):
    encontrados = False
    for col in range(colunaBrick, colunaBrick + max_splits):
        valor = ws.cell(row=row, column=col).value
        if valor and valor in baseBrick:
            encontrados = True
            break
    if not encontrados:
        nome_cliente = ws.cell(row=row, column=cabecalhos["Lista de clientes-alvo: Nome"]).value or ""
        nome_alvo = ws.cell(row=row, column=cabecalhos["Alvo: Alvos"]).value or ""
        nome_conta = ws.cell(row=row, column=cabecalhos["Nome da conta"]).value or ""
        ciclo_marketing = ws.cell(row=row, column=cabecalhos["Ciclo de Marketing"]).value or ""

        cliente_partes = nome_cliente.split()
        alvo_partes = nome_alvo.split()

        primeiro_cliente = cliente_partes[0] if cliente_partes else ""
        primeiro_alvo = alvo_partes[0] if alvo_partes else ""
        sobrenomes_alvo = " ".join(alvo_partes[1:]) if len(alvo_partes) > 1 else ""

        ws_Limpeza.append([
            primeiro_cliente, primeiro_alvo, sobrenomes_alvo,
            nome_conta, 0, ciclo_marketing
        ])

# Etapa 7: Salvar e mover para pasta do ciclo
os.makedirs(pasta_base, exist_ok=True)
nome_arquivo_novo = f'DELETE TCLs- {setorRep}.xlsx'
caminho_arquivo_novo = os.path.join(pasta_base, nome_arquivo_novo)
wb.save(caminho_arquivo_novo)

data_criacao = datetime.fromtimestamp(os.path.getctime(caminho_arquivo_novo))
nome_ciclo = obter_ciclo_por_data(data_criacao)

if nome_ciclo:
    pasta_ciclo = os.path.join(pasta_base, nome_ciclo)
    os.makedirs(pasta_ciclo, exist_ok=True)
    caminho_final = os.path.join(pasta_ciclo, nome_arquivo_novo)
    shutil.move(caminho_arquivo_novo, caminho_final)
    print(f"✅ Arquivo salvo e movido para: {caminho_final}")
else:
    print("⚠️ Nenhum ciclo correspondente encontrado.")

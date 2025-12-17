import os
import re
import shutil
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

# === CONFIGURAÇÕES ===
pasta_base = r'C:\Users\pazlimx1\OneDrive - Abbott\Documents\AUTOMACAO\ADICAO PAINEL'
arquivo_origem = r'C:\Users\pazlimx1\OneDrive - Abbott\Documents\AUTOMACAO\ADD TCL\BASE\BASE_ADD.xlsx'

CICLOS = [
    {"nome": "CICLO 07", "inicio": "2025-07-18", "fim": "2025-08-15"},
    {"nome": "CICLO 08", "inicio": "2025-08-18", "fim": "2025-09-15"},
    {"nome": "CICLO 09", "inicio": "2025-09-16", "fim": "2025-10-14"},
    {"nome": "CICLO 10", "inicio": "2025-10-15", "fim": "2025-11-12"},
    {"nome": "CICLO 11", "inicio": "2025-11-13", "fim": "2025-12-17"},
]
for ciclo in CICLOS:
    ciclo["inicio"] = datetime.strptime(ciclo["inicio"], "%Y-%m-%d")
    ciclo["fim"] = datetime.strptime(ciclo["fim"], "%Y-%m-%d")

# === FUNÇÕES ===

def obter_ciclo_por_data(data):
    for ciclo in CICLOS:
        if ciclo["inicio"] <= data <= ciclo["fim"]:
            return ciclo["nome"]
    return None

def formatar_bricks(codigos):
    def format(codigo):
        codigo = codigo.strip()
        return f'BR_{codigo.zfill(7)}' if not codigo.startswith("BR_") else codigo
    return [format(c) for c in codigos if c]

def separar_bricks(ws, colunaBrick):
    max_splits = 1
    for row in range(2, ws.max_row + 1):
        brick = ws.cell(row=row, column=colunaBrick).value
        if brick and isinstance(brick, str):
            partes = brick.split()
            max_splits = max(max_splits, len(partes))
    for _ in range(max_splits - 1):
        ws.insert_cols(colunaBrick + 1)
    for row in range(2, ws.max_row + 1):
        brick = ws.cell(row=row, column=colunaBrick).value
        if brick and isinstance(brick, str):
            partes = brick.split()
            for i, parte in enumerate(partes):
                ws.cell(row=row, column=colunaBrick + i, value=parte)
    return max_splits

def aplicar_vlookup(ws, colunaBrick, colunaFinal, baseBrick):
    for i, codigo in enumerate(baseBrick, start=2):
        ws.cell(row=i, column=colunaFinal + 1, value=codigo)
    for row in range(2, ws.max_row + 1):
        letra_g = get_column_letter(colunaBrick)
        letra_usuario = get_column_letter(colunaFinal + 1)
        letra_vazia = get_column_letter(colunaFinal)
        formula = f'=IFERROR(VLOOKUP({letra_g}{row},{letra_usuario}:{letra_usuario},1,FALSE),"")'
        ws[f'{letra_vazia}{row}'] = formula

def criar_aba_adicao(ws, wb, baseBrick, colunaBrick, max_splits):
    if "ADICAO" not in wb.sheetnames:
        ws_Limpeza = wb.create_sheet("ADICAO")
    else:
        ws_Limpeza = wb["ADICAO"]
        ws_Limpeza.delete_rows(1, ws_Limpeza.max_row)

    cabecalhos = {
        "Ciclo de Marketing": None,
        "Alvo: Território": None,
        "Account ID_18": None,
        "Nome da conta": None,
        "Specialty 1": None,
        "Contact ID_18": None,
        "Licença Médica Legal": None
    }

    for col in range(1, ws.max_column + 1):
        valor = ws.cell(row=1, column=col).value
        if valor in cabecalhos:
            cabecalhos[valor] = col

    ws_Limpeza.append([
        "Ciclo de Marketing:", "Alvo: Território:",
        "Account ID_18:", "Nome da conta:", "Specialty 1:",
        "Contact ID_18:", "Licença Médica Legal"
    ])

    ids_unicos = set()
    linhas_filtradas = []

    for row in range(2, ws.max_row + 1):
        encontrados = any(
            ws.cell(row=row, column=col).value in baseBrick
            for col in range(colunaBrick, colunaBrick + max_splits)
        )
        if encontrados:
            account_id = ws.cell(row=row, column=cabecalhos["Account ID_18"]).value or ""
            if account_id not in ids_unicos:
                ids_unicos.add(account_id)
                linha = [
                    ws.cell(row=row, column=cabecalhos["Ciclo de Marketing"]).value or "",
                    ws.cell(row=row, column=cabecalhos["Alvo: Território"]).value or "",
                    account_id,
                    ws.cell(row=row, column=cabecalhos["Nome da conta"]).value or "",
                    ws.cell(row=row, column=cabecalhos["Specialty 1"]).value or "",
                    ws.cell(row=row, column=cabecalhos["Contact ID_18"]).value or "",
                    ws.cell(row=row, column=cabecalhos["Licença Médica Legal"]).value or ""
                ]
                linhas_filtradas.append(linha)

    for linha in linhas_filtradas:
        ws_Limpeza.append(linha)

def gerar_nome_unico(pasta, nome_base):
    nome, ext = os.path.splitext(nome_base)
    contador = 1
    nome_final = nome_base
    while os.path.exists(os.path.join(pasta, nome_final)):
        contador += 1
        nome_final = f"{nome}_v{contador}{ext}"
    return nome_final

def salvar_e_mover_arquivo(wb, setorRep, pasta_base):
    os.makedirs(pasta_base, exist_ok=True)
    nome_arquivo_base = f'ADD TCLs- {setorRep}.xlsx'
    nome_arquivo_novo = gerar_nome_unico(pasta_base, nome_arquivo_base)
    caminho_arquivo_novo = os.path.join(pasta_base, nome_arquivo_novo)

    try:
        wb.save(caminho_arquivo_novo)
    except PermissionError:
        print("❌ Não foi possível salvar o arquivo. Feche o Excel e tente novamente.")
        return

    data_criacao = datetime.fromtimestamp(os.path.getctime(caminho_arquivo_novo))
    nome_ciclo = obter_ciclo_por_data(data_criacao)

    if nome_ciclo:
        pasta_ciclo = os.path.join(pasta_base, nome_ciclo)
        os.makedirs(pasta_ciclo, exist_ok=True)
        caminho_final = os.path.join(pasta_ciclo, nome_arquivo_novo)

        if os.path.exists(caminho_final):
            try:
                os.remove(caminho_final)
            except PermissionError:
                print("❌ O arquivo de destino está aberto. Feche-o e tente novamente.")
                return

        try:
            shutil.move(caminho_arquivo_novo, caminho_final)
            print(f"✅ Arquivo salvo e movido para: {caminho_final}")
        except PermissionError:
            print("❌ Não foi possível mover o arquivo. Feche o Excel e tente novamente.")
    else:
        print("⚠️ Nenhum ciclo correspondente encontrado.")

# === EXECUÇÃO PRINCIPAL ===

def main():
    setorRep = input("Informe o Setor do Representante: ")
    baseBrick_input = input("Cole os códigos BRICK separados por espaço, vírgula ou quebra de linha:\n")
    codigosRaw = re.split(r'[,\s]+', baseBrick_input.strip())
    baseBrick = formatar_bricks(codigosRaw)

    wb = load_workbook(arquivo_origem)
    ws = wb.active
    colunaBrick = 7
    max_splits = separar_bricks(ws, colunaBrick)
    colunaFinal = colunaBrick + max_splits
    ws.insert_cols(colunaFinal)
    ws.insert_cols(colunaFinal + 1)

    aplicar_vlookup(ws, colunaBrick, colunaFinal, baseBrick)
    criar_aba_adicao(ws, wb, baseBrick, colunaBrick, max_splits)
    salvar_e_mover_arquivo(wb, setorRep, pasta_base)

if __name__ == "__main__":
    main()

